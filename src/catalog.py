from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field as PydanticField, field_validator
from rewire import simple_plugin
from rewire_sqlmodel import session_context, transaction

from src.models import Category, Product

plugin = simple_plugin()

REQUIRED_COLUMNS = {
    "sku",
    "name",
    "category",
    "retail_price",
    "wholesale_price",
}


class CatalogValidationError(ValueError):
    pass


class CatalogRow(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    sku: str
    name: str
    category: str
    category_image_url: Optional[str] = None
    description: str = ""
    characteristics: str = ""
    retail_price: Decimal = PydanticField(gt=0)
    wholesale_price: Decimal = PydanticField(ge=0)
    discount_price: Optional[Decimal] = PydanticField(default=None, gt=0)
    image_url: Optional[str] = None
    is_active: bool = True
    is_popular: bool = False
    is_recommended: bool = False
    owner: str = "Булат"
    owner_share_percent: Decimal = PydanticField(default=Decimal("70"), ge=0, le=100)

    @field_validator("discount_price")
    @classmethod
    def discount_must_be_lower(cls, value: Optional[Decimal], info):
        retail_price = info.data.get("retail_price")
        if value is not None and retail_price is not None and value >= retail_price:
            raise ValueError("discount_price must be lower than retail_price")
        return value


@dataclass(frozen=True)
class SyncReport:
    created: int = 0
    updated: int = 0
    hidden: int = 0
    categories_created: int = 0


def _as_bool(value: Any, default: bool) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "y", "да", "активен"}:
        return True
    if normalized in {"0", "false", "no", "n", "нет", "скрыт"}:
        return False
    raise ValueError(f"cannot parse boolean value {value!r}")


def _as_decimal(value: Any) -> Optional[Decimal]:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(" ", "").replace(",", "."))
    except InvalidOperation as exc:
        raise ValueError(f"cannot parse money value {value!r}") from exc


def load_catalog_rows(path: str | Path) -> list[CatalogRow]:
    source = Path(path)
    if not source.is_file():
        raise CatalogValidationError(f"Catalog file not found: {source}")

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if "products" not in workbook.sheetnames:
            raise CatalogValidationError("Workbook must contain a 'products' sheet")

        sheet = workbook["products"]
        raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not raw_headers:
            raise CatalogValidationError("The products sheet is empty")

        headers = [str(value).strip() if value is not None else "" for value in raw_headers]
        missing = REQUIRED_COLUMNS - set(headers)
        if missing:
            raise CatalogValidationError(
                f"Missing required columns: {', '.join(sorted(missing))}"
            )

        rows: list[CatalogRow] = []
        seen_skus: set[str] = set()
        errors: list[str] = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            raw = dict(zip(headers, values, strict=False))
            if not any(value not in (None, "") for value in raw.values()):
                continue
            try:
                raw.update(
                    retail_price=_as_decimal(raw.get("retail_price")),
                    wholesale_price=_as_decimal(raw.get("wholesale_price")),
                    discount_price=_as_decimal(raw.get("discount_price")),
                    is_active=_as_bool(raw.get("is_active"), True),
                    is_popular=_as_bool(raw.get("is_popular"), False),
                    is_recommended=_as_bool(raw.get("is_recommended"), False),
                    owner_share_percent=_as_decimal(raw.get("owner_share_percent"))
                                        or Decimal("70"),
                )
                parsed = CatalogRow.model_validate(raw)
                normalized_sku = parsed.sku.casefold()
                if normalized_sku in seen_skus:
                    raise ValueError(f"duplicate sku {parsed.sku!r}")
                seen_skus.add(normalized_sku)
                rows.append(parsed)
            except Exception as exc:
                errors.append(f"row {row_number}: {exc}")

        if errors:
            raise CatalogValidationError("Invalid catalog:\n" + "\n".join(errors))
        if not rows:
            raise CatalogValidationError("The products sheet contains no products")
        return rows
    finally:
        workbook.close()


@transaction(1)
async def sync_catalog(path: str | Path) -> SyncReport:
    rows = load_catalog_rows(path)
    session = session_context.get()
    now = datetime.now()

    categories = {category.name.casefold(): category for category in await Category.select().all()}
    products = {product.sku.casefold(): product for product in await Product.select().all()}
    source_skus = {row.sku.casefold() for row in rows}

    created = updated = hidden = categories_created = 0
    for row in rows:
        category_key = row.category.casefold()
        category = categories.get(category_key)
        if not category:
            category = Category(name=row.category, image_url=row.category_image_url).add()
            await session.flush()
            categories[category_key] = category
            categories_created += 1
        else:
            category.is_active = True
            if row.category_image_url:
                category.image_url = row.category_image_url

            category.updated_at = now
            category.add()

        product = products.get(row.sku.casefold())
        values = row.model_dump(
            exclude={"category", "category_image_url"},
            exclude_none=False
        )

        values["category_id"] = category.id
        values["updated_at"] = now

        if not product:
            Product(**values).add()
            created += 1
        else:
            product.sqlmodel_update(values)
            product.add()
            updated += 1

    for sku, product in products.items():
        if sku not in source_skus and product.is_active:
            product.is_active = False
            product.updated_at = now
            product.add()
            hidden += 1

    return SyncReport(
        created=created,
        updated=updated,
        hidden=hidden,
        categories_created=categories_created,
    )
